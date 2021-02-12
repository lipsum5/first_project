import openpyxl as px
import numpy as np

book = px.Workbook()
sheet = book.active

x = np.random.rand(100)
x_list = x.tolist()

k = 0

for i in range(5):
    col = []
    for j in range(20):
        sheet.cell(row = j+1, column = i+1, value = x_list[k])
        col.append(x_list[k])
        k += 1
    ave = sum(col) / len(col)
    sheet.cell(row = 23, column = i+1, value = ave)

book.save('sample.xlsx')
